import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('data/taxonomy.json', 'r') as f:
    data = json.load(f)

niches = sorted(data['niches'], key=lambda n: (n['tier'], n['name']))

wb = Workbook()

# --- Sheet 1: Keyword Overview ---
ws1 = wb.active
ws1.title = 'Keyword Overview'

header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
header_fill = PatternFill('solid', fgColor='1a1a1a')
title_font = Font(bold=True, size=16, name='Arial', color='FFFFFF')
title_fill = PatternFill('solid', fgColor='000000')
body_font = Font(size=10, name='Arial')
alt_fill = PatternFill('solid', fgColor='f5f5f5')
thin_border = Border(
    bottom=Side(style='thin', color='dddddd')
)

headers = ['Tier', 'Niche', 'Primary Keywords', 'Long-Tail Subtopics', 'Target Audience', 'Pain Points', 'Content Strategy', 'Monetization']

# Title row
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
title_cell = ws1.cell(row=1, column=1, value='Incognito Browser — Keyword Research')
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 40

# Subtitle
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
sub_cell = ws1.cell(row=2, column=1, value=f'44 niches | {sum(len(n["keywords"]) for n in niches)} primary keywords | {sum(len(n["context"]["subtopics"]) for n in niches)} long-tail subtopics')
sub_cell.font = Font(size=10, name='Arial', color='888888')
sub_cell.fill = PatternFill('solid', fgColor='0a0a0a')
ws1.row_dimensions[2].height = 25

# Header row
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws1.row_dimensions[3].height = 30

# Data rows
for i, niche in enumerate(niches):
    row = i + 4
    ctx = niche.get('context', {})
    values = [
        f'Tier {niche["tier"]}',
        niche['name'],
        ', '.join(niche['keywords']),
        ', '.join(ctx.get('subtopics', [])),
        ctx.get('audience', ''),
        ctx.get('pain_points', ''),
        ctx.get('content_that_works', ''),
        ctx.get('monetization', ''),
    ]
    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = alt_fill

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 32
ws1.column_dimensions['C'].width = 45
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 40
ws1.column_dimensions['F'].width = 45
ws1.column_dimensions['G'].width = 40
ws1.column_dimensions['H'].width = 40

ws1.freeze_panes = 'A4'

# --- Sheet 2: Full Keyword List ---
ws2 = wb.create_sheet('Full Keyword List')

headers2 = ['Keyword', 'Type', 'Parent Niche', 'Tier']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[1].height = 30

rows = []
for niche in niches:
    for kw in niche['keywords']:
        rows.append((kw, 'Primary', niche['name'], niche['tier']))
    for st in niche.get('context', {}).get('subtopics', []):
        rows.append((st, 'Subtopic', niche['name'], niche['tier']))

rows.sort(key=lambda r: r[0].lower())

primary_fill = PatternFill('solid', fgColor='e8f0fe')
subtopic_fill = PatternFill('solid', fgColor='fef7e0')

for i, (kw, kw_type, parent, tier) in enumerate(rows):
    row = i + 2
    ws2.cell(row=row, column=1, value=kw).font = body_font
    type_cell = ws2.cell(row=row, column=2, value=kw_type)
    type_cell.font = body_font
    ws2.cell(row=row, column=3, value=parent).font = body_font
    ws2.cell(row=row, column=4, value=f'Tier {tier}').font = body_font
    for col in range(1, 5):
        ws2.cell(row=row, column=col).border = thin_border
        if kw_type == 'Primary':
            ws2.cell(row=row, column=col).fill = primary_fill
        elif i % 2 == 1:
            ws2.cell(row=row, column=col).fill = alt_fill

ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 10
ws2.freeze_panes = 'A2'

output_path = 'Incognito Browser - Keyword Research.xlsx'
wb.save(output_path)
print(f'Created {output_path} with {len(niches)} niches and {len(rows)} total keywords')
